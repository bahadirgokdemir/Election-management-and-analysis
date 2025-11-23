"""
Export servisi - Excel ve PDF formatında veri aktarımı
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
from django.http import HttpResponse
from django.db.models import QuerySet, Count
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """
    Kişi verilerini farklı formatlarda export etmek için servis
    """

    COLUMN_DEFINITIONS = {
        'kisi_sicilno': {'label': 'Sicil No', 'width': 15, 'align': 'center'},
        'ad': {'label': 'Ad', 'width': 20, 'align': 'left'},
        'soyad': {'label': 'Soyad', 'width': 20, 'align': 'left'},
        'mail': {'label': 'E-posta', 'width': 30, 'align': 'left'},
        'telno': {'label': 'Telefon', 'width': 18, 'align': 'center'},
        'ilce': {'label': 'İlçe', 'width': 18, 'align': 'left'},
        'adres_aciklama': {'label': 'Adres Açıklama', 'width': 35, 'align': 'left'},
        'notlar': {'label': 'Notlar', 'width': 40, 'align': 'left'},
        'cevap_status': {'label': 'Cevap Durumu', 'width': 18, 'align': 'center'},
        'avukat': {'label': 'Avukat', 'width': 25, 'align': 'left'},
    }

    @staticmethod
    def extract_row_data(lp, selected_columns: List[str]) -> List[str]:
        """LawyerPerson objesinden seçili kolonları extract eder"""
        row = []
        for col in selected_columns:
            if col == 'kisi_sicilno':
                row.append(lp.kisi_sicilno or '')
            elif col == 'ad':
                row.append(lp.ad or '')
            elif col == 'soyad':
                row.append(lp.soyad or '')
            elif col == 'mail':
                row.append(lp.mail or '')
            elif col == 'telno':
                row.append(lp.telno or '')
            elif col == 'ilce':
                row.append(lp.ilce or '')
            elif col == 'adres_aciklama':
                row.append(lp.adres_aciklama or '')
            elif col == 'notlar':
                row.append(lp.notlar or '')
            elif col == 'cevap_status':
                row.append(lp.cevap_status.label if lp.cevap_status else '')
            elif col == 'avukat':
                row.append(f"{lp.lawyer.ad} {lp.lawyer.soyad} ({lp.lawyer.sicil_no})" if lp.lawyer else '')
        return row

    @staticmethod
    def export_to_csv(queryset: QuerySet, selected_columns: List[str]) -> HttpResponse:
        """CSV formatında export"""
        import csv

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="kisiler_export_{timestamp}.csv"'

        writer = csv.writer(response)

        # Header
        headers = [ExportService.COLUMN_DEFINITIONS[col]['label'] for col in selected_columns]
        writer.writerow(headers)

        # Data
        for lp in queryset:
            row = ExportService.extract_row_data(lp, selected_columns)
            writer.writerow(row)

        return response

    @staticmethod
    def export_to_excel(
        queryset: QuerySet,
        selected_columns: List[str],
        include_stats: bool = True,
        include_filters: Optional[Dict[str, Any]] = None
    ) -> HttpResponse:
        """
        Excel formatında gelişmiş export
        - Formatlanmış header
        - Otomatik sütun genişlikleri
        - Stil ve renkler
        - İsteğe bağlı istatistikler
        - Filtre bilgileri
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Kişiler"

        # Stil tanımlamaları
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        info_font = Font(name='Calibri', size=9, italic=True, color='666666')
        info_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        data_font = Font(name='Calibri', size=10)
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        current_row = 1

        # Başlık ve bilgi satırları
        if include_stats or include_filters:
            # Tarih ve saat
            ws.merge_cells(f'A{current_row}:{get_column_letter(len(selected_columns))}{current_row}')
            date_cell = ws[f'A{current_row}']
            date_cell.value = f'Export Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
            date_cell.font = info_font
            date_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # İstatistikler
            if include_stats:
                total_count = queryset.count()

                # Durum bazında sayım
                status_counts = queryset.values('cevap_status__label').annotate(count=Count('id')).order_by('-count')

                # İlçe bazında sayım
                district_counts = queryset.exclude(ilce__isnull=True).exclude(ilce='').values('ilce').annotate(count=Count('id')).order_by('-count')[:5]

                ws.merge_cells(f'A{current_row}:{get_column_letter(len(selected_columns))}{current_row}')
                stats_cell = ws[f'A{current_row}']
                stats_cell.value = f'Toplam Kayıt: {total_count}'
                stats_cell.font = Font(name='Calibri', size=10, bold=True, color='333333')
                stats_cell.fill = info_fill
                stats_cell.alignment = Alignment(horizontal='center')
                current_row += 1

            # Filtre bilgileri
            if include_filters:
                filter_texts = []
                if include_filters.get('q'):
                    filter_texts.append(f"Arama: {include_filters['q']}")
                if include_filters.get('status_label'):
                    filter_texts.append(f"Durum: {include_filters['status_label']}")
                if include_filters.get('lawyer_name'):
                    filter_texts.append(f"Avukat: {include_filters['lawyer_name']}")
                if include_filters.get('ilce'):
                    filter_texts.append(f"İlçe: {include_filters['ilce']}")

                if filter_texts:
                    ws.merge_cells(f'A{current_row}:{get_column_letter(len(selected_columns))}{current_row}')
                    filter_cell = ws[f'A{current_row}']
                    filter_cell.value = '🔍 Filtreler: ' + ' | '.join(filter_texts)
                    filter_cell.font = info_font
                    filter_cell.fill = info_fill
                    filter_cell.alignment = Alignment(horizontal='left')
                    current_row += 1

            # Boş satır
            current_row += 1

        # Header row (sütun başlıkları)
        for col_idx, col_key in enumerate(selected_columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = ExportService.COLUMN_DEFINITIONS[col_key]['label']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

            # Sütun genişliğini ayarla
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = ExportService.COLUMN_DEFINITIONS[col_key]['width']

        current_row += 1

        # Data rows
        for lp in queryset:
            row_data = ExportService.extract_row_data(lp, selected_columns)

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.font = data_font
                cell.border = border

                # Alignment
                col_key = selected_columns[col_idx - 1]
                align_type = ExportService.COLUMN_DEFINITIONS[col_key]['align']
                cell.alignment = Alignment(horizontal=align_type, vertical='top', wrap_text=True)

            current_row += 1

        # Freeze panes (başlık satırını dondur)
        if include_stats or include_filters:
            ws.freeze_panes = ws.cell(row=current_row - queryset.count(), column=1)
        else:
            ws.freeze_panes = 'A2'

        # Auto-filter ekle
        ws.auto_filter.ref = ws.dimensions

        # BytesIO'ya kaydet
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        # Response oluştur
        response = HttpResponse(
            bio.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="kisiler_export_{timestamp}.xlsx"'

        return response

    @staticmethod
    def export_to_pdf(
        queryset: QuerySet,
        selected_columns: List[str],
        include_stats: bool = True,
        include_filters: Optional[Dict[str, Any]] = None
    ) -> HttpResponse:
        """
        PDF formatında export
        ReportLab kullanarak profesyonel PDF çıktısı
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            # ReportLab yüklü değilse, kullanıcıya bilgi ver
            response = HttpResponse(content_type='text/plain')
            response['Content-Disposition'] = 'attachment; filename="error.txt"'
            response.write('PDF export için "reportlab" kütüphanesi gereklidir.\n\n')
            response.write('Kurulum için: pip install reportlab')
            return response

        # PDF buffer
        buffer = BytesIO()

        # Landscape orientation (daha fazla sütun için)
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        # Stil tanımları
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=1  # Center
        )
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=1
        )

        # Sayfa elemanları
        elements = []

        # Tarih
        date_text = f'Export Tarihi: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}'
        elements.append(Paragraph(date_text, info_style))
        elements.append(Spacer(1, 0.5*cm))

        # İstatistikler
        if include_stats:
            stats_text = f'<b>Toplam Kayıt:</b> {queryset.count()}'
            elements.append(Paragraph(stats_text, info_style))
            elements.append(Spacer(1, 0.3*cm))

        # Filtre bilgileri
        if include_filters:
            filter_texts = []
            if include_filters.get('q'):
                filter_texts.append(f"Arama: {include_filters['q']}")
            if include_filters.get('status_label'):
                filter_texts.append(f"Durum: {include_filters['status_label']}")
            if include_filters.get('lawyer_name'):
                filter_texts.append(f"Avukat: {include_filters['lawyer_name']}")
            if include_filters.get('ilce'):
                filter_texts.append(f"İlçe: {include_filters['ilce']}")

            if filter_texts:
                filter_text = '🔍 <b>Filtreler:</b> ' + ' | '.join(filter_texts)
                elements.append(Paragraph(filter_text, info_style))
                elements.append(Spacer(1, 0.5*cm))

        # Tablo verileri
        table_data = []

        # Header
        headers = [ExportService.COLUMN_DEFINITIONS[col]['label'] for col in selected_columns]
        table_data.append(headers)

        # Data rows
        for lp in queryset[:500]:  # PDF için max 500 satır (performans)
            row_data = ExportService.extract_row_data(lp, selected_columns)
            # Uzun metinleri kısalt
            row_data = [str(cell)[:50] + '...' if len(str(cell)) > 50 else str(cell) for cell in row_data]
            table_data.append(row_data)

        # Tablo oluştur
        # Sütun genişliklerini hesapla
        available_width = landscape(A4)[0] - 2*cm
        col_widths = []
        total_width_units = sum(ExportService.COLUMN_DEFINITIONS[col]['width'] for col in selected_columns)
        for col in selected_columns:
            width_ratio = ExportService.COLUMN_DEFINITIONS[col]['width'] / total_width_units
            col_widths.append(available_width * width_ratio)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        # Tablo stili
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),

            # Data
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),

            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),

            # Alternate row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
        ]))

        elements.append(table)

        # Uyarı (500+ kayıt varsa)
        if queryset.count() > 500:
            elements.append(Spacer(1, 0.5*cm))
            warning_style = ParagraphStyle(
                'Warning',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.red,
                alignment=1
            )
            warning_text = f'⚠️ Performans nedeniyle sadece ilk 500 kayıt gösterilmektedir. (Toplam: {queryset.count()})'
            elements.append(Paragraph(warning_text, warning_style))

        # PDF oluştur
        doc.build(elements)

        # Response
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename="kisiler_export_{timestamp}.pdf"'

        return response
