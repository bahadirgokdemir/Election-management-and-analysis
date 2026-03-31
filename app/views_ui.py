from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models import Q
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, JsonResponse
from django.db import transaction
import csv
from io import BytesIO
from datetime import date as date_obj

from .models import Lawyer, Person, StatusOption, LawyerPerson, UploadBatch, UploadRowStaging, Election, BaroLawyer, BaroLawyerTag, CommitteeMembership
from .permissions import login_required_custom, admin_required, uploader_required
from .services.importer import parse_and_stage
from .services.diff_service import compute_diff
from .services.apply_service import apply_diff
from .services.reports import report_overview
from .services.unique_people_service import UniquePeopleService
from .services.person_analytics_service import PersonAnalyticsService
from .services.baro_loader import BaroLoader
from .utils.search import apply_name_search

from django.shortcuts import get_object_or_404


@login_required_custom
@require_http_methods(["GET"])
def ui_dashboard(request):
    try:
        data = report_overview()
    except Exception as e:
        data = {}
        messages.error(request, f'Dashboard verisi yüklenirken hata oluştu: {e}')
    return render(request, 'app/dashboard.html', {'data': data})


@login_required_custom
@csrf_exempt
@require_http_methods(["GET", "POST"])
def ui_lawyers(request):
    # POST icin admin yetkisi gerekli
    if request.method == "POST":
        if not request.user.is_superuser:
            try:
                if not request.user.profile.is_admin():
                    messages.error(request, 'Bu islem icin admin yetkisi gerekiyor.')
                    return redirect('ui_lawyers')
            except:
                messages.error(request, 'Bu islem icin admin yetkisi gerekiyor.')
                return redirect('ui_lawyers')
        sicil = (request.POST.get('sicil') or '').strip()
        ad = (request.POST.get('ad') or '').strip()
        soyad = (request.POST.get('soyad') or '').strip()

        if not sicil or not ad or not soyad:
            messages.error(request, "Sicil No, Ad ve Soyad zorunludur.")
            return redirect('ui_lawyers')

        try:
            lawyer, created = Lawyer.objects.get_or_create(
                sicil_no=sicil,
                defaults={'ad': ad, 'soyad': soyad}
            )
        except Exception as e:
            messages.error(request, f'Avukat kaydedilirken hata oluştu: {e}')
            return redirect('ui_lawyers')
        if created:
            messages.success(request, f"Avukat eklendi: {sicil} — {ad} {soyad}")
        else:
            if lawyer.ad != ad or lawyer.soyad != soyad:
                messages.info(request, f"Bu sicil zaten kayıtlı: {lawyer.sicil_no} — {lawyer.ad} {lawyer.soyad}. "
                                       f"Girdiğiniz isim uygulanmadı.")
            else:
                messages.info(request, "Bu sicil no zaten mevcut; yeni bir kayıt oluşturulmadı.")
        return redirect('ui_lawyers')

    _SYS_SICILS = [_KARA_LISTE_SICIL, _BEYAZ_LISTE_SICIL]
    q = (request.GET.get('q') or '').strip()
    qs = Lawyer.objects.exclude(sicil_no__in=_SYS_SICILS).order_by('-id')
    if q:
        qs = apply_name_search(qs, q, extra_fields=['sicil_no'])
    page = Paginator(qs, 20).get_page(request.GET.get('page'))
    return render(request, 'app/lawyers_list.html', {'page': page, 'q': q})


@login_required_custom
@require_http_methods(["GET"])
def ui_people(request):
    q = (request.GET.get('q') or '').strip()
    status_key = request.GET.get('status')
    lawyer_id = request.GET.get('lawyer')
    selected_ilce = request.GET.get('ilce')

    # Gelişmiş arama parametreleri
    adv_sicil = (request.GET.get('sicil') or '').strip()
    adv_ad = (request.GET.get('ad') or '').strip()
    adv_soyad = (request.GET.get('soyad') or '').strip()
    adv_mail = (request.GET.get('mail') or '').strip()
    adv_telno = (request.GET.get('telno') or '').strip()
    adv_ilce = (request.GET.get('ilce_search') or '').strip()
    adv_adres = (request.GET.get('adres') or '').strip()
    adv_notlar = (request.GET.get('notlar') or '').strip()
    # Baro kayıt alanları (BaroLawyer'den)
    adv_dogum_yeri = (request.GET.get('dogum_yeri') or '').strip()
    adv_nufus_il = (request.GET.get('nufus_il') or '').strip()

    # LawyerPerson ilişkilerini getir - her ilişki ayrı satır olacak
    # KRITIK: Artık tüm veriler LawyerPerson'da, Person'a bakmıyoruz
    qs = LawyerPerson.objects.select_related(
        'cevap_status',
        'lawyer'
    ).filter(active=True).order_by('-id', 'lawyer__ad', 'lawyer__soyad')

    # Genel arama (tüm alanlarda) - Türkçe karakter ve ad+soyad destekli
    if q:
        qs = apply_name_search(qs, q, extra_fields=[
            'kisi_sicilno', 'mail', 'ilce', 'telno', 'adres_aciklama', 'notlar'
        ])

    # Alan bazında gelişmiş aramalar - LawyerPerson alanları (Türkçe karakter destekli)
    from app.utils.search import normalize_tr, _normalize_db_expr
    if adv_sicil:
        qs = qs.filter(kisi_sicilno__icontains=adv_sicil)
    if adv_ad:
        qs = qs.annotate(_adv_ad=_normalize_db_expr('ad')).filter(_adv_ad__icontains=normalize_tr(adv_ad))
    if adv_soyad:
        qs = qs.annotate(_adv_soyad=_normalize_db_expr('soyad')).filter(_adv_soyad__icontains=normalize_tr(adv_soyad))
    if adv_mail:
        qs = qs.filter(mail__icontains=adv_mail)
    if adv_telno:
        qs = qs.filter(telno__icontains=adv_telno)
    if adv_ilce:
        qs = qs.filter(ilce__icontains=adv_ilce)
    if adv_adres:
        qs = qs.filter(adres_aciklama__icontains=adv_adres)
    if adv_notlar:
        qs = qs.filter(notlar__icontains=adv_notlar)
    # Baro kayıt alanlarına göre filtrele (kisi_sicilno üzerinden BaroLawyer'e join)
    if adv_dogum_yeri:
        matching = BaroLawyer.objects.filter(dogum_yeri__icontains=adv_dogum_yeri).values_list('sicil_no', flat=True)
        qs = qs.filter(kisi_sicilno__in=matching)
    if adv_nufus_il:
        matching = BaroLawyer.objects.filter(nufus_il__icontains=adv_nufus_il).values_list('sicil_no', flat=True)
        qs = qs.filter(kisi_sicilno__in=matching)

    # İlçe dropdown filtresi
    if selected_ilce and selected_ilce != 'None':
        qs = qs.filter(ilce=selected_ilce)

    # Durum filtresi
    if status_key and status_key != 'None':
        qs = qs.filter(cevap_status__key=status_key)

    # Avukat filtresi
    if lawyer_id and lawyer_id != 'None':
        qs = qs.filter(lawyer_id=lawyer_id)

    # İlçe listesi (dropdown için) - LawyerPerson'dan al
    districts = LawyerPerson.objects.filter(active=True).exclude(ilce__isnull=True).exclude(ilce='').values_list('ilce', flat=True).distinct().order_by('ilce')

    page = Paginator(qs, 25).get_page(request.GET.get('page'))
    statuses = StatusOption.objects.all().order_by('key')
    lawyers = Lawyer.objects.exclude(sicil_no__in=[_KARA_LISTE_SICIL, _BEYAZ_LISTE_SICIL]).order_by('ad', 'soyad')

    return render(request, 'app/people_list.html', {
        'page': page,
        'q': q,
        'status_key': status_key,
        'lawyer_id': lawyer_id,
        'selected_ilce': selected_ilce,
        'districts': districts,
        'statuses': statuses,
        'lawyers': lawyers,
        'adv_dogum_yeri': adv_dogum_yeri,
        'adv_nufus_il': adv_nufus_il,
    })


@uploader_required
@csrf_exempt
@require_http_methods(["GET", "POST"])
def ui_upload(request):
    # Aktif seçim kontrolü
    active_election = Election.objects.filter(is_active=True).first()
    if active_election:
        messages.error(request, f'Aktif seçim devam ediyor ({active_election.name}). Seçim bitene kadar yükleme yapamazsınız.')
        return redirect('ui_dashboard')

    if request.method == 'GET':
        # Tüm avukatları alfabetik sırala (sistem avukatları hariç)
        lawyers = Lawyer.objects.exclude(sicil_no__in=[_KARA_LISTE_SICIL, _BEYAZ_LISTE_SICIL]).order_by('ad', 'soyad')
        return render(request, 'app/upload_wizard.html', {'lawyers': lawyers})

    # POST
    file = request.FILES.get('file')
    if not file:
        messages.error(request, 'Dosya zorunludur.')
        return redirect('ui_upload')

    # Avukat seçimi: Mevcut mi yoksa yeni mi?
    existing_lawyer_id = request.POST.get('existing_lawyer_id')

    if existing_lawyer_id:
        # Mevcut avukat seçildi
        try:
            lawyer = Lawyer.objects.get(id=existing_lawyer_id)
            messages.info(request, f'Seçilen avukat: {lawyer.ad} {lawyer.soyad} ({lawyer.sicil_no})')
        except Lawyer.DoesNotExist:
            messages.error(request, 'Seçilen avukat bulunamadı.')
            return redirect('ui_upload')
    else:
        # Yeni avukat ekleniyor
        sicil_no = (request.POST.get('lawyer_sicil') or '').strip()
        ad = (request.POST.get('lawyer_ad') or '').strip()
        soyad = (request.POST.get('lawyer_soyad') or '').strip()

        if not sicil_no or not ad or not soyad:
            messages.error(request, 'Yeni avukat için Sicil No, Ad ve Soyad zorunludur.')
            return redirect('ui_upload')

        lawyer, created = Lawyer.objects.get_or_create(
            sicil_no=sicil_no,
            defaults={'ad': ad, 'soyad': soyad}
        )
        if created:
            messages.success(request, f'Yeni avukat oluşturuldu: {sicil_no} — {ad} {soyad}')
        else:
            if (lawyer.ad != ad) or (lawyer.soyad != soyad):
                messages.info(
                    request,
                    f"Bu sicil no zaten kayıtlı: {sicil_no} — {lawyer.ad} {lawyer.soyad}. "
                    f"Gönderdiğiniz isim ({ad} {soyad}) kayda uygulanmadı."
                )
            else:
                messages.info(request, f'Mevcut avukat kullanılıyor: {lawyer.ad} {lawyer.soyad}')

    try:
        # 1) Dosyayı staging'e yükle
        from app.utils.file_validators import ValidationError

        try:
            batch_id, row_count, name_mismatches = parse_and_stage(
                file, lawyer.id,
                created_by=str(request.user) if request.user.is_authenticated else None
            )
        except ValidationError as ve:
            # Detaylı validasyon hatası
            error_msg = f'Dosya Validasyon Hatası: {ve.message}'
            messages.error(request, error_msg)

            # Detayları ayrı mesajlar olarak ekle
            if ve.details:
                for detail in ve.details[:10]:  # Max 10 detay göster
                    messages.warning(request, f'  • {detail}')

            return redirect('ui_upload')

        # Batch notlarını kontrol et (Baro uyarıları için)
        batch = UploadBatch.objects.get(id=batch_id)
        if batch.notes:
            messages.info(request, f'{batch.notes}')

        # İsim uyuşmazlığı uyarıları
        if name_mismatches:
            messages.warning(
                request,
                f'{len(name_mismatches)} kişi Baro kaydındaki isimle uyuşmadığı için listeye eklenmedi:'
            )
            for m in name_mismatches[:15]:  # Max 15 detay göster
                messages.warning(
                    request,
                    f'Sicil {m["kisi_sicilno"]} (Satır {m["row_num"]}): '
                    f'Listede "{m["excel_ad"]} {m["excel_soyad"]}" '
                    f'— Baro\'da "{m["baro_ad"]} {m["baro_soyad"]}"'
                )
            if len(name_mismatches) > 15:
                messages.warning(request, f'... ve {len(name_mismatches) - 15} kişi daha.')

        # 2) Otomatik olarak uygula — merge_mode=True: eski kayıtları silme
        actor = str(request.user) if request.user.is_authenticated else None
        result = apply_diff(batch_id, actor=actor, merge_mode=True)

        if result.get('ok'):
            counts = result.get('counts', {})
            added = counts.get('added', 0)
            changed = counts.get('changed', 0)
            would_remove = counts.get('would_remove', 0)
            rejection_note = f', {len(name_mismatches)} kişi isim uyuşmazlığı nedeniyle reddedildi' if name_mismatches else ''
            messages.success(
                request,
                f'Yükleme başarılı! {row_count} satır işlendi. '
                f'{added} yeni kayıt eklendi, {changed} kayıt güncellendi{rejection_note}.'
            )
            if would_remove:
                messages.warning(
                    request,
                    f'{would_remove} kayıt yeni listede yer almıyor ancak korundu (silinmedi). '
                    f'Bu kayıtları silmek için ilgili kişiyi manuel olarak pasifleştirebilirsiniz.'
                )
        else:
            messages.warning(request, f'Yükleme tamamlandı ancak uygulama sırasında sorun oluştu: {result.get("message")}')

        # Baro'da bulunmayan kayıt var mı? Varsa inceleme sayfasına yönlendir
        batch_sicils = list(UploadRowStaging.objects.filter(batch_id=batch_id).values_list('kisi_sicilno', flat=True))
        if batch_sicils:
            existing_baro_sicils = set(
                BaroLawyer.objects.filter(sicil_no__in=batch_sicils).values_list('sicil_no', flat=True)
            )
            unknown_count = sum(1 for s in batch_sicils if s not in existing_baro_sicils)
            if unknown_count:
                messages.info(
                    request,
                    f'{unknown_count} kişi Baro veritabanında bulunamadı. '
                    f'Aşağıdan inceleyip sisteme ekleyebilirsiniz.'
                )
                return redirect('ui_upload_unknown_records', batch_id=batch_id)

        return redirect('ui_dashboard')
    except Exception as e:
        messages.error(request, f'Beklenmeyen hata: {e}')
        return redirect('ui_upload')


@uploader_required
@require_http_methods(["GET"])
def ui_diff_preview(request, batch_id: int):
    try:
        diff = compute_diff(batch_id)
    except Exception as e:
        messages.error(request, f'Diff önizlemesi yüklenirken hata oluştu: {e}')
        return redirect('ui_dashboard')
    return render(request, 'app/diff_preview.html', {'diff': diff})


@uploader_required
@csrf_exempt
@require_http_methods(["POST"])
def ui_approve_batch(request, batch_id: int):
    try:
        res = apply_diff(batch_id, actor=str(request.user) if request.user.is_authenticated else None)
    except Exception as e:
        messages.error(request, f'Değişiklikler uygulanırken hata oluştu: {e}')
        return redirect('ui_diff_preview', batch_id=batch_id)
    if res.get('ok'):
        messages.success(request, 'Değişiklikler uygulandı.')
        return redirect('ui_dashboard')
    messages.error(request, res.get('message', 'Uygulama başarısız.'))
    return redirect('ui_diff_preview', batch_id=batch_id)


@uploader_required
@require_http_methods(["GET", "POST"])
def ui_upload_unknown_records(request, batch_id: int):
    """
    Yüklenen listede BaroLawyer tablosunda bulunmayan kayıtları göster.
    Admin bunları tek tek veya toplu olarak BaroLawyer'a ekleyebilir.
    """
    batch = get_object_or_404(UploadBatch, id=batch_id)

    staging_rows = list(UploadRowStaging.objects.filter(batch_id=batch_id))
    all_sicils = [r.kisi_sicilno for r in staging_rows]
    existing_baro_sicils = set(
        BaroLawyer.objects.filter(sicil_no__in=all_sicils).values_list('sicil_no', flat=True)
    )
    unknown_rows = [r for r in staging_rows if r.kisi_sicilno not in existing_baro_sicils]

    if request.method == 'POST':
        action = request.POST.get('action', '')
        selected_sicils = set(request.POST.getlist('sicil_nos'))

        if action == 'add_all':
            to_add = unknown_rows
        elif action == 'add_selected':
            to_add = [r for r in unknown_rows if r.kisi_sicilno in selected_sicils]
        else:
            to_add = []

        added_count = 0
        skipped_count = 0
        for row in to_add:
            if BaroLawyer.objects.filter(sicil_no=row.kisi_sicilno).exists():
                skipped_count += 1
                continue
            ad = row.ad if row.ad and row.ad != 'Bilinmiyor' else ''
            soyad = row.soyad if row.soyad and row.soyad != 'Bilinmiyor' else ''
            BaroLawyer.objects.create(
                sicil_no=row.kisi_sicilno,
                ad=ad,
                soyad=soyad,
                tel=row.telno or '',
                mail=row.mail or '',
            )
            added_count += 1

        if added_count:
            messages.success(request, f'{added_count} kayıt Baro veritabanına eklendi.')
        if skipped_count:
            messages.info(request, f'{skipped_count} kayıt zaten Baro veritabanında bulunduğundan atlandı.')

        return redirect('ui_dashboard')

    return render(request, 'app/unknown_baro_records.html', {
        'batch': batch,
        'unknown_rows': unknown_rows,
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_people_export_preview(request):
    """Export önizlemesi için ilk 10 satırı ve kullanılabilir sütunları döndürür."""
    q = (request.GET.get('q') or '').strip()
    status_key = request.GET.get('status')
    lawyer_id = request.GET.get('lawyer')

    # LawyerPerson kayıtlarını çek - her kayıt ayrı satır
    qs = LawyerPerson.objects.select_related('cevap_status', 'lawyer').filter(active=True)

    if q:
        qs = apply_name_search(qs, q, extra_fields=[
            'kisi_sicilno', 'mail', 'ilce', 'telno', 'adres_aciklama', 'notlar'
        ])
    if status_key and status_key not in ('None', '', 'null'):
        qs = qs.filter(cevap_status__key=status_key)
    if lawyer_id and lawyer_id not in ('None', '', 'null'):
        try:
            qs = qs.filter(lawyer_id=int(lawyer_id))
        except (ValueError, TypeError):
            pass

    # İlçe filtresi ekle
    ilce_filter = request.GET.get('ilce')
    if ilce_filter and ilce_filter not in ('None', '', 'null'):
        qs = qs.filter(ilce=ilce_filter)

    # Kullanılabilir sütunlar
    available_columns = [
        {'key': 'kisi_sicilno', 'label': 'Sicil No', 'default': True},
        {'key': 'ad', 'label': 'Ad', 'default': True},
        {'key': 'soyad', 'label': 'Soyad', 'default': True},
        {'key': 'mail', 'label': 'E-posta', 'default': True},
        {'key': 'telno', 'label': 'Telefon', 'default': False},
        {'key': 'ilce', 'label': 'İlçe', 'default': True},
        {'key': 'adres_aciklama', 'label': 'Adres Açıklama', 'default': False},
        {'key': 'notlar', 'label': 'Notlar', 'default': False},
        {'key': 'cevap_status', 'label': 'Cevap Durumu', 'default': True},
        {'key': 'avukat', 'label': 'Avukat', 'default': True},
    ]

    # İlk 10 satır önizleme
    preview_data = []
    for lp in qs[:10]:
        preview_data.append({
            'kisi_sicilno': lp.kisi_sicilno,
            'ad': lp.ad,
            'soyad': lp.soyad,
            'mail': lp.mail or '',
            'telno': lp.telno or '',
            'ilce': lp.ilce or '',
            'adres_aciklama': lp.adres_aciklama or '',
            'notlar': lp.notlar or '',
            'cevap_status': lp.cevap_status.label if lp.cevap_status else '',
            'avukat': f"{lp.lawyer.ad} {lp.lawyer.soyad}" if lp.lawyer else '',
        })

    # Filtre seçenekleri için listeler
    districts = list(LawyerPerson.objects.filter(active=True).exclude(ilce__isnull=True).exclude(ilce='').values_list('ilce', flat=True).distinct().order_by('ilce'))
    statuses = list(StatusOption.objects.all().values('key', 'label'))
    lawyers = list(Lawyer.objects.all().values('id', 'ad', 'soyad').order_by('ad', 'soyad'))

    return JsonResponse({
        'columns': available_columns,
        'preview': preview_data,
        'total_count': qs.count(),
        'districts': districts,
        'statuses': statuses,
        'lawyers': lawyers,
    })


@login_required_custom
@csrf_exempt
@require_http_methods(["POST"])
def ui_people_export_download(request):
    """
    Seçilen sütunlarla filtrelenmiş kişileri istenen formatta (CSV/Excel/PDF) indirir.
    """
    from .services.export_service import ExportService

    q = (request.POST.get('q') or '').strip()
    status_key = request.POST.get('status')
    lawyer_id = request.POST.get('lawyer')
    selected_ilce = request.POST.get('ilce')
    selected_columns = request.POST.getlist('columns[]')
    export_format = request.POST.get('format', 'csv')  # csv, excel, pdf
    include_stats = request.POST.get('include_stats', 'true') == 'true'

    if not selected_columns:
        return JsonResponse({'error': 'Lütfen en az bir sütun seçin'}, status=400)

    # LawyerPerson kayıtlarını çek - her kayıt ayrı satır
    qs = LawyerPerson.objects.select_related('cevap_status', 'lawyer').filter(active=True).order_by('id')

    # Filtreler
    if q:
        qs = apply_name_search(qs, q, extra_fields=[
            'kisi_sicilno', 'mail', 'ilce', 'telno', 'adres_aciklama', 'notlar'
        ])
    if status_key and status_key not in ('None', '', 'null'):
        qs = qs.filter(cevap_status__key=status_key)
    if lawyer_id and lawyer_id not in ('None', '', 'null'):
        try:
            qs = qs.filter(lawyer_id=int(lawyer_id))
        except (ValueError, TypeError):
            pass
    if selected_ilce and selected_ilce not in ('None', '', 'null'):
        qs = qs.filter(ilce=selected_ilce)

    # Filtre bilgileri (Excel/PDF için)
    filter_info = {}
    if q:
        filter_info['q'] = q
    if status_key and status_key not in ('None', '', 'null'):
        status = StatusOption.objects.filter(key=status_key).first()
        if status:
            filter_info['status_label'] = status.label
    if lawyer_id and lawyer_id not in ('None', '', 'null'):
        try:
            lawyer = Lawyer.objects.filter(id=int(lawyer_id)).first()
            if lawyer:
                filter_info['lawyer_name'] = f"{lawyer.ad} {lawyer.soyad}"
        except (ValueError, TypeError):
            pass
    if selected_ilce and selected_ilce not in ('None', '', 'null'):
        filter_info['ilce'] = selected_ilce

    # Format'a göre export
    try:
        if export_format == 'excel':
            return ExportService.export_to_excel(
                queryset=qs,
                selected_columns=selected_columns,
                include_stats=include_stats,
                include_filters=filter_info
            )
        elif export_format == 'pdf':
            return ExportService.export_to_pdf(
                queryset=qs,
                selected_columns=selected_columns,
                include_stats=include_stats,
                include_filters=filter_info
            )
        else:  # csv (default)
            return ExportService.export_to_csv(
                queryset=qs,
                selected_columns=selected_columns
            )
    except Exception as e:
        return JsonResponse({'error': f'Dışa aktarma sırasında hata oluştu: {e}'}, status=500)


@login_required_custom
@require_http_methods(["GET"])
def ui_people_export(request):
    """Filtrelenmiş kişileri CSV olarak indirir (eski yöntem - geriye dönük uyumluluk)."""
    q = (request.GET.get('q') or '').strip()
    status_key = request.GET.get('status')
    lawyer_id = request.GET.get('lawyer')

    qs = Person.objects.select_related('cevap_status').prefetch_related(
        'lawyerperson_set__lawyer'
    ).all()

    if q:
        qs = qs.filter(
            Q(kisi_sicilno__icontains=q) |
            Q(ad__icontains=q) |
            Q(soyad__icontains=q) |
            Q(mail__icontains=q) |
            Q(ilce__icontains=q) |
            Q(telno__icontains=q) |
            Q(adres_aciklama__icontains=q) |
            Q(notlar__icontains=q)
        )
    if status_key and status_key != 'None':
        qs = qs.filter(cevap_status__key=status_key)
    if lawyer_id and lawyer_id != 'None':
        qs = qs.filter(lawyerperson__lawyer_id=lawyer_id)

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="kisiler.csv"'

    writer = csv.writer(response)
    writer.writerow(['kisi_sicilno', 'ad', 'soyad', 'mail', 'ilce', 'cevap_status', 'avukatlar'])
    for p in qs:
        # Get all lawyers for this person
        lawyers_str = ', '.join([f"{lp.lawyer.ad} {lp.lawyer.soyad} ({lp.lawyer.sicil_no})"
                                 for lp in p.lawyerperson_set.all()])
        writer.writerow([
            p.kisi_sicilno,
            p.ad,
            p.soyad,
            p.mail or '',
            p.ilce or '',
            p.cevap_status.key if p.cevap_status else '',
            lawyers_str
        ])

    return response


# Sablon indirme - CSV
@login_required_custom
@require_http_methods(["GET"])
def ui_download_template_csv(request):
    """
    Excel/CSV şablonu (kolon başlıkları):
    sicilno,ad,soyad,cevapDurumu,telno,mail,ilce,adres_aciklama,notlar

    NOT: Excel'de eksik olan bilgiler Baro kayıtlarından otomatik tamamlanır.
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="liste_sablon.csv"'
    writer = csv.writer(response)
    writer.writerow(['sicilno', 'ad', 'soyad', 'cevapDurumu', 'telno', 'mail', 'ilce', 'adres_aciklama', 'notlar'])
    return response


# Sablon indirme - XLSX (openpyxl ile)
@login_required_custom
@require_http_methods(["GET"])
def ui_download_template_xlsx(request):
    """
    Excel şablonu: Tüm kolonlar mevcut
    Excel'de eksik olan bilgiler Baro kayıtlarından otomatik tamamlanır.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except Exception:
        # openpyxl yoksa CSV döndür
        return ui_download_template_csv(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Seçmen Listesi"

    # Minimal modern renk paleti
    BRAND_BLUE = "2563eb"         # Modern mavi
    BG_HEADER = "f8fafc"          # Çok açık gri
    BG_INPUT = "ffffff"           # Beyaz
    TEXT_PRIMARY = "0f172a"       # Koyu metin
    TEXT_SECONDARY = "64748b"     # Gri metin
    BORDER_COLOR = "e2e8f0"       # Açık border

    # Minimal border
    clean_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    # === GÖNDERİCİ BİLGİLERİ ===
    ws.merge_cells('A1:E1')
    sender_header = ws['A1']
    sender_header.value = 'Gönderici Bilgileri'
    sender_header.font = Font(name='Calibri', size=10, bold=True, color=TEXT_SECONDARY)
    sender_header.fill = PatternFill(start_color=BG_HEADER, end_color=BG_HEADER, fill_type="solid")
    sender_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    sender_header.border = clean_border
    ws.row_dimensions[1].height = 24

    # Kolon genişlikleri
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 26
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 26
    ws.column_dimensions['I'].width = 30

    # Satır 2: Input alanları (doğrudan)
    placeholders = ['12345', 'Ahmet', 'Yılmaz', '0532 123 4567', 'ornek@email.com']
    for col_idx, placeholder in enumerate(placeholders, start=1):
        cell = ws.cell(row=2, column=col_idx, value=placeholder)
        cell.font = Font(name='Calibri', size=10, italic=True, color=TEXT_SECONDARY)
        cell.fill = PatternFill(start_color=BG_INPUT, end_color=BG_INPUT, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = clean_border

    ws.row_dimensions[2].height = 28

    # Satır 3: Boş ayırıcı
    ws.row_dimensions[3].height = 8

    # === LİSTE BÖLÜMÜ ===
    ws.merge_cells('A4:I4')
    list_header = ws['A4']
    list_header.value = 'Seçmen Listesi'
    list_header.font = Font(name='Calibri', size=10, bold=True, color=TEXT_SECONDARY)
    list_header.fill = PatternFill(start_color=BG_HEADER, end_color=BG_HEADER, fill_type="solid")
    list_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    list_header.border = clean_border
    ws.row_dimensions[4].height = 24

    # Satır 5: Tablo başlıkları
    data_headers = ['Sicil No', 'Ad', 'Soyad', 'Cevap Durumu', 'Telefon', 'E-posta', 'İlçe', 'Adres', 'Notlar']
    for col_idx, header in enumerate(data_headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=9, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = clean_border

    ws.row_dimensions[5].height = 26

    # === CEVAP DURUMU DROPDOWN ===
    dv = DataValidation(
        type="list",
        formula1='"OLUMLU,OLUMSUZ,TEPKİLİ,İLETİŞİM GEREKİYOR"',
        allow_blank=True,
        showErrorMessage=True,
        showInputMessage=True
    )
    dv.error = 'Sadece listeden seçim yapabilirsiniz'
    dv.errorTitle = 'Geçersiz Giriş'
    dv.errorStyle = 'stop'
    dv.prompt = 'Lütfen bir seçenek seçin'
    dv.promptTitle = 'Cevap Durumu'

    # D sütununun 6. satırından itibaren
    dv.add('D6:D1000')
    ws.add_data_validation(dv)

    # === AÇIKLAMA KUTUSU ===
    ws.column_dimensions['K'].width = 2.5  # Ayırıcı
    ws.column_dimensions['L'].width = 38   # Açıklama

    # Başlık
    info_header = ws.cell(row=1, column=12, value='Cevap Seçenekleri Açıklaması')
    info_header.font = Font(name='Calibri', size=9, bold=True, color=TEXT_PRIMARY)
    info_header.fill = PatternFill(start_color=BG_HEADER, end_color=BG_HEADER, fill_type="solid")
    info_header.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    info_header.border = clean_border

    # Açıklamalar
    explanations = [
        'OLUMLU: Veysel Kırıcı\'yı destekliyor ve ön seçime katılacak',
        'OLUMSUZ: Başka bir adayı destekliyor',
        'TEPKİLİ: Ön seçime veya baroya karşı olumsuz tutum',
        'İLETİŞİM GEREKİYOR: Ziyaret edilmeli, fikri yok veya aranmalı',
    ]

    for idx, text in enumerate(explanations, start=2):
        cell = ws.cell(row=idx, column=12, value=text)
        cell.font = Font(name='Calibri', size=8, color=TEXT_PRIMARY)
        cell.fill = PatternFill(start_color=BG_INPUT, end_color=BG_INPUT, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
        cell.border = clean_border

    # Satır yükseklikleri
    ws.row_dimensions[1].height = 22
    for i in range(2, 6):
        ws.row_dimensions[i].height = 20

    # Açıklama sayfası ekle
    ws_info = wb.create_sheet("Bilgi")
    ws_info.append(['ÖNEMLİ BİLGİLENDİRME'])
    ws_info.append([])
    ws_info.append(['Excel\'de eksik bıraktığınız bilgiler (telefon, e-posta, adres)'])
    ws_info.append(['Baro kayıtlarından OTOMATIK olarak tamamlanacaktır.'])
    ws_info.append([])
    ws_info.append(['Gerekli Kolonlar: sicilno, ad, soyad'])
    ws_info.append(['Opsiyonel Kolonlar: cevapDurumu, telno, mail, ilce, adres_aciklama, notlar'])
    ws_info.append([])
    ws_info.append(['Sicil numaraları Baro kayıtlarında kontrol edilir.'])
    ws_info.append(['Baro\'da olmayan siciller için UYARI mesajı gösterilir.'])

    # Bilgi sayfası başlığını stillendir
    ws_info['A1'].font = Font(bold=True, size=14, color="C00000")

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = 'attachment; filename="liste_sablon.xlsx"'
    return resp
# ========== YENİ: Seçili satırları uygula ==========
@uploader_required
@csrf_exempt
@require_http_methods(["POST"])
def ui_approve_selected(request, batch_id: int):
    """
    Diff'ten sadece seçilen satırları uygular.
    Form inputları:
      - added:   çoklu checkbox (value=kisi_sicilno)
      - removed: çoklu checkbox (value=kisi_sicilno)
      - changed: çoklu checkbox (value=kisi_sicilno) -> tüm değişen alanlar uygulanır
    """
    try:
        diff = compute_diff(batch_id)
    except Exception as e:
        messages.error(request, f'Diff bilgisi alınırken hata oluştu: {e}')
        return redirect('ui_dashboard')

    sel_added = set(request.POST.getlist('added'))
    sel_removed = set(request.POST.getlist('removed'))
    sel_changed = set(request.POST.getlist('changed'))

    # diff.lawyer.* bilgileri diff objesinde mevcut
    lawyer_id = diff['lawyer']['id'] if isinstance(diff.get('lawyer'), dict) and diff['lawyer'].get('id') else None
    if not lawyer_id:
        messages.error(request, 'Avukat bilgisi bulunamadı.')
        return redirect('ui_diff_preview', batch_id=batch_id)

    try:
        lawyer = Lawyer.objects.get(id=lawyer_id)
    except Lawyer.DoesNotExist:
        messages.error(request, 'Avukat kaydı bulunamadı.')
        return redirect('ui_diff_preview', batch_id=batch_id)

    # Yardımcı: status key -> instance
    def get_status(key):
        if not key:
            return None
        try:
            return StatusOption.objects.get(key=key)
        except StatusOption.DoesNotExist:
            return None

    applied_add, applied_remove, applied_change = 0, 0, 0

    try:
        with transaction.atomic():
            # ADDED - Yeni kayıtlar
            for row in diff.get('added', []):
                ks = str(row.get('kisi_sicilno') or '')
                if ks not in sel_added:
                    continue

                # Person global kaydı (referans için) - sadece sicilno saklanır
                person, _ = Person.objects.get_or_create(
                    kisi_sicilno=ks,
                    defaults={
                        'ad': row.get('ad') or '',
                        'soyad': row.get('soyad') or '',
                    }
                )

                # KRITIK: Gerçek veriler LawyerPerson'da saklanır
                # Her avukat için bağımsız kopya
                LawyerPerson.objects.update_or_create(
                    lawyer=lawyer,
                    kisi_sicilno=ks,
                    defaults={
                        'person': person,
                        'ad': row.get('ad') or '',
                        'soyad': row.get('soyad') or '',
                        'mail': row.get('mail') or '',
                        'telno': row.get('telno') or '',
                        'ilce': row.get('ilce') or '',
                        'adres_aciklama': row.get('adres_aciklama') or '',
                        'notlar': row.get('notlar') or '',
                        'cevap_status': get_status(row.get('cevap_status_key')),
                        'active': True
                    }
                )
                applied_add += 1

            # REMOVED - Kayıt silme (soft delete)
            for row in diff.get('removed', []):
                ks = str(row.get('kisi_sicilno') or '')
                if ks not in sel_removed:
                    continue

                # Bu avukattan soft-delete et (active=False) - kayıtları koru
                updated_count = LawyerPerson.objects.filter(
                    lawyer=lawyer,
                    kisi_sicilno=ks
                ).update(active=False)

                if updated_count > 0:
                    applied_remove += 1

            # CHANGED - Mevcut kayıtları güncelle
            for row in diff.get('changed', []):
                ks = str(row.get('kisi_sicilno') or '')
                if ks not in sel_changed:
                    continue

                after = row.get('after') or {}

                # Person referansını al/oluştur
                person, _ = Person.objects.get_or_create(
                    kisi_sicilno=ks,
                    defaults={
                        'ad': after.get('ad') or '',
                        'soyad': after.get('soyad') or '',
                    }
                )

                # KRITIK: LawyerPerson'ı güncelle - sadece bu avukatın verisi
                LawyerPerson.objects.update_or_create(
                    lawyer=lawyer,
                    kisi_sicilno=ks,
                    defaults={
                        'person': person,
                        'ad': after.get('ad') or '',
                        'soyad': after.get('soyad') or '',
                        'mail': after.get('mail') or '',
                        'telno': after.get('telno') or '',
                        'ilce': after.get('ilce') or '',
                        'adres_aciklama': after.get('adres_aciklama') or '',
                        'notlar': after.get('notlar') or '',
                        'cevap_status': get_status(after.get('cevap_status_key')),
                        'active': True
                    }
                )
                applied_change += 1
    except Exception as e:
        messages.error(request, f'Değişiklikler uygulanırken hata oluştu: {e}')
        return redirect('ui_diff_preview', batch_id=batch_id)

    messages.success(
        request,
        f"Seçili değişiklikler uygulandı. (+{applied_add} / -{applied_remove} / Δ{applied_change})"
    )
    return redirect('ui_dashboard')


@login_required_custom
@require_http_methods(["GET"])
def ui_lawyer_people(request, lawyer_id: int):
    lawyer = get_object_or_404(Lawyer, id=lawyer_id)
    q = (request.GET.get('q') or '').strip()
    status_key = request.GET.get('status')

    qs = Person.objects.filter(
        lawyerperson__lawyer=lawyer
    ).select_related('cevap_status').order_by('soyad', 'ad')

    if q:
        qs = qs.filter(
            Q(kisi_sicilno__icontains=q) |
            Q(ad__icontains=q) |
            Q(soyad__icontains=q) |
            Q(mail__icontains=q) |
            Q(ilce__icontains=q) |
            Q(telno__icontains=q) |
            Q(adres_aciklama__icontains=q) |
            Q(notlar__icontains=q)
        )
    if status_key:
        qs = qs.filter(cevap_status__key=status_key)

    page = Paginator(qs, 25).get_page(request.GET.get('page'))
    statuses = StatusOption.objects.all().order_by('key')

    return render(request, 'app/lawyer_people.html', {
        'lawyer': lawyer,
        'page': page,
        'q': q,
        'status_key': status_key,
        'statuses': statuses,
    })


@login_required_custom
@require_http_methods(["GET", "POST"])
def ui_person_edit(request, person_id):
    """
    LawyerPerson kaydını düzenle.
    ÖNEMLI: person_id aslında LawyerPerson ID'sidir.
    Her avukatın kendi verisi var, Person tablosu sadece referans.
    """
    lp = get_object_or_404(LawyerPerson, id=person_id)

    if request.method == "GET":
        # LawyerPerson bilgilerini JSON olarak döndür
        statuses = StatusOption.objects.all().order_by('key')
        return JsonResponse({
            'id': lp.id,
            'kisi_sicilno': lp.kisi_sicilno,
            'ad': lp.ad,
            'soyad': lp.soyad,
            'mail': lp.mail or '',
            'telno': lp.telno or '',
            'ilce': lp.ilce or '',
            'adres_aciklama': lp.adres_aciklama or '',
            'notlar': lp.notlar or '',
            'cevap_status_key': lp.cevap_status.key if lp.cevap_status else '',
            'available_statuses': [{'key': s.key, 'label': s.label} for s in statuses]
        })

    if request.method == "POST":
        # LawyerPerson bilgilerini güncelle - sadece bu avukat için
        import json
        try:
            data = json.loads(request.body)
        except (json.JSONDecodeError, Exception) as e:
            return JsonResponse({'success': False, 'error': f'Geçersiz veri formatı: {e}'}, status=400)

        try:
            lp.ad = data.get('ad', lp.ad)
            lp.soyad = data.get('soyad', lp.soyad)
            lp.mail = data.get('mail') or None
            lp.telno = data.get('telno') or None
            lp.ilce = data.get('ilce') or None
            lp.adres_aciklama = data.get('adres_aciklama') or None
            lp.notlar = data.get('notlar') or None

            status_key = data.get('cevap_status_key')
            if status_key:
                lp.cevap_status = StatusOption.objects.filter(key=status_key).first()
            else:
                lp.cevap_status = None

            lp.save()
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Kayıt güncellenirken hata oluştu: {e}'}, status=500)
        return JsonResponse({'success': True})


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
def ui_person_relation_delete(request, lawyerperson_id):
    """LawyerPerson ilişkisini sil (soft delete - active=False)"""
    # Aktif seçim kontrolü
    active_election = Election.objects.filter(is_active=True).first()
    if active_election:
        return JsonResponse({'success': False, 'error': f'Aktif seçim devam ediyor ({active_election.name}). Seçim bitene kadar silme işlemi yapamazsınız.'})

    lp = get_object_or_404(LawyerPerson, id=lawyerperson_id)
    lp.active = False
    lp.save()
    return JsonResponse({'success': True})


@admin_required
@csrf_exempt
@require_http_methods(["POST"])
def ui_lawyer_delete(request, lawyer_id):
    """Avukatı ve ona ait tüm ilişkileri sil"""
    try:
        # Aktif seçim kontrolü
        active_election = Election.objects.filter(is_active=True).first()
        if active_election:
            return JsonResponse({
                'success': False,
                'error': f'Aktif seçim devam ediyor ({active_election.name}). Seçim bitene kadar silme işlemi yapamazsınız.'
            })

        lawyer = get_object_or_404(Lawyer, id=lawyer_id)

        # Avukata ait kişi sayısını al
        person_count = LawyerPerson.objects.filter(lawyer=lawyer, active=True).count()

        with transaction.atomic():
            # Tüm ilişkileri sil
            LawyerPerson.objects.filter(lawyer=lawyer).delete()

            # Yükleme kayıtlarını sil
            UploadBatch.objects.filter(lawyer=lawyer).delete()

            # Avukatı sil
            lawyer.delete()

        return JsonResponse({
            'success': True,
            'deleted_relations': person_count
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Silme işlemi sırasında bir hata oluştu: {str(e)}'
        })


@login_required_custom
@require_http_methods(["GET"])
def ui_unique_people(request):
    """
    Benzersiz kişiler sayfası - Tekrarlı kayıtlar birleştirilmiş
    Her sicil no için tek kayıt, ama tüm avukat ve bilgiler birleştirilmiş
    """
    q = (request.GET.get('q') or '').strip()
    status_key = request.GET.get('status')
    lawyer_id = request.GET.get('lawyer')
    selected_ilce = request.GET.get('ilce')
    min_records = request.GET.get('min_records')  # Tekrarlı kayıtları filtreleme

    # Min records filtresi
    min_records_int = None
    if min_records and min_records.isdigit():
        min_records_int = int(min_records)

    # Benzersiz kişileri getir
    unique_people = UniquePeopleService.get_unique_people(
        search_query=q,
        status_key=status_key if status_key and status_key != 'None' else None,
        lawyer_id=int(lawyer_id) if lawyer_id and lawyer_id != 'None' else None,
        district=selected_ilce if selected_ilce and selected_ilce != 'None' else None,
        min_records=min_records_int
    )

    # İstatistikler
    stats = UniquePeopleService.get_statistics()

    # Pagination
    from django.core.paginator import Paginator
    page = Paginator(unique_people, 25).get_page(request.GET.get('page'))

    # Filtre seçenekleri
    statuses = StatusOption.objects.all().order_by('key')
    lawyers = Lawyer.objects.all().order_by('ad', 'soyad')
    districts = LawyerPerson.objects.filter(active=True).exclude(ilce__isnull=True).exclude(ilce='').values_list('ilce', flat=True).distinct().order_by('ilce')

    return render(request, 'app/unique_people.html', {
        'page': page,
        'q': q,
        'status_key': status_key,
        'lawyer_id': lawyer_id,
        'selected_ilce': selected_ilce,
        'min_records': min_records,
        'districts': districts,
        'statuses': statuses,
        'lawyers': lawyers,
        'stats': stats,
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_unique_person_detail(request, kisi_sicilno: str):
    """Belirli bir sicil no için detay modal verisi"""
    try:
        person = UniquePeopleService.get_person_details(kisi_sicilno)
    except Exception as e:
        return JsonResponse({'error': f'Kişi detayı alınırken hata oluştu: {e}'}, status=500)

    if not person:
        return JsonResponse({'error': 'Kişi bulunamadı'}, status=404)

    return JsonResponse(person)


@login_required_custom
@require_http_methods(["GET"])
def ui_person_analytics(request, kisi_sicilno: str):
    """
    Kişi bazlı analiz verisi
    Avukat ve durum istatistikleri, grafik verileri
    """
    try:
        analytics = PersonAnalyticsService.get_person_analytics(kisi_sicilno)
    except Exception as e:
        return JsonResponse({'error': f'Analiz verisi alınırken hata oluştu: {e}'}, status=500)

    if not analytics:
        return JsonResponse({'error': 'Kişi bulunamadı veya veri yok'}, status=404)

    try:
        comparison = PersonAnalyticsService.get_comparison_stats(kisi_sicilno)
        analytics['comparison'] = comparison
    except Exception:
        analytics['comparison'] = {}

    return JsonResponse(analytics)


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_lawyers(request):
    """
    Database'deki tüm baro kayıtlarını listeler - Gelişmiş filtreleme ve sıralama ile
    """
    q = (request.GET.get('q') or '').strip()

    # Gelişmiş arama parametreleri
    adv_sicil = (request.GET.get('sicil') or '').strip()
    adv_ad = (request.GET.get('ad') or '').strip()
    adv_soyad = (request.GET.get('soyad') or '').strip()
    adv_mail = (request.GET.get('mail') or '').strip()
    adv_tel = (request.GET.get('tel') or '').strip()
    adv_adres = (request.GET.get('adres') or '').strip()
    adv_dogum_yeri = (request.GET.get('dogum_yeri') or '').strip()
    adv_nufus_il = (request.GET.get('nufus_il') or '').strip()
    adv_cinsiyet = (request.GET.get('cinsiyet') or '').strip()
    adv_ilce = (request.GET.get('ilce') or '').strip()

    # Sicil aralığı parametreleri
    sicil_min = request.GET.get('sicil_min', '').strip()
    sicil_max = request.GET.get('sicil_max', '').strip()

    # Sıralama parametresi
    order_by = request.GET.get('order_by', '').strip()

    # Tag filtresi
    tag_filter = request.GET.get('tag_filter', '').strip()  # 'blacklist', 'whitelist', 'none', ''

    # Database'den kayıtları çek (tag bilgisiyle birlikte)
    from django.db.models import Count as _Count
    qs = BaroLawyer.objects.select_related('tag').annotate(membership_count=_Count('memberships')).all()

    # Genel arama filtresi - Türkçe karakter ve ad+soyad destekli
    if q:
        qs = apply_name_search(qs, q, extra_fields=['sicil_no', 'mail', 'tel', 'adres'])

    # Alan bazında gelişmiş aramalar (Türkçe karakter destekli ad/soyad)
    from app.utils.search import normalize_tr, _normalize_db_expr as _ndb
    if adv_sicil:
        qs = qs.filter(sicil_no__icontains=adv_sicil)
    if adv_ad:
        qs = qs.annotate(_baro_adv_ad=_ndb('ad')).filter(_baro_adv_ad__icontains=normalize_tr(adv_ad))
    if adv_soyad:
        qs = qs.annotate(_baro_adv_soyad=_ndb('soyad')).filter(_baro_adv_soyad__icontains=normalize_tr(adv_soyad))
    if adv_mail:
        qs = qs.filter(mail__icontains=adv_mail)
    if adv_tel:
        qs = qs.filter(tel__icontains=adv_tel)
    if adv_adres:
        qs = qs.filter(adres__icontains=adv_adres)
    if adv_dogum_yeri:
        qs = qs.filter(dogum_yeri__icontains=adv_dogum_yeri)
    if adv_nufus_il:
        qs = qs.filter(nufus_il__icontains=adv_nufus_il)
    if adv_cinsiyet:
        qs = qs.filter(cinsiyet__icontains=adv_cinsiyet)
    if adv_ilce:
        qs = qs.filter(ilce__icontains=adv_ilce)

    # Sicil aralığı filtresi ve sıralama için import
    from django.db.models.functions import Cast
    from django.db.models import IntegerField

    # Sicil aralığı filtresi - Numeric comparison
    if sicil_min and sicil_min.isdigit():
        qs = qs.annotate(sicil_int_filter=Cast('sicil_no', IntegerField())).filter(sicil_int_filter__gte=int(sicil_min))
    if sicil_max and sicil_max.isdigit():
        if not sicil_min or not sicil_min.isdigit():
            qs = qs.annotate(sicil_int_filter=Cast('sicil_no', IntegerField()))
        qs = qs.filter(sicil_int_filter__lte=int(sicil_max))

    # Sıralama - Sicil no'yu numeric olarak sırala
    allowed_orders = ['sicil_no', '-sicil_no']
    if order_by in allowed_orders:
        # Sicil no'yu integer'a cast ederek sırala
        if order_by == 'sicil_no':
            qs = qs.annotate(sicil_int=Cast('sicil_no', IntegerField())).order_by('sicil_int')
        else:  # -sicil_no
            qs = qs.annotate(sicil_int=Cast('sicil_no', IntegerField())).order_by('-sicil_int')
    else:
        # Varsayılan sıralama: sicil no artan (numeric)
        qs = qs.annotate(sicil_int=Cast('sicil_no', IntegerField())).order_by('sicil_int')
        order_by = 'sicil_no'

    # Tag filtresi uygula
    if tag_filter == 'blacklist':
        qs = qs.filter(tag__tag_type='blacklist')
    elif tag_filter == 'whitelist':
        qs = qs.filter(tag__tag_type='whitelist')
    elif tag_filter == 'none':
        qs = qs.filter(tag__isnull=True)

    # İstatistikler
    total = BaroLawyer.objects.count()
    blacklist_count = BaroLawyerTag.objects.filter(tag_type='blacklist').count()
    whitelist_count = BaroLawyerTag.objects.filter(tag_type='whitelist').count()
    stats = {
        'total_records': total,
        'blacklist_count': blacklist_count,
        'whitelist_count': whitelist_count,
        'untagged_count': total - blacklist_count - whitelist_count,
    }

    # Pagination
    page = Paginator(qs, 25).get_page(request.GET.get('page'))

    return render(request, 'app/baro_lawyers.html', {
        'page': page,
        'q': q,
        'adv_sicil': adv_sicil,
        'adv_ad': adv_ad,
        'adv_soyad': adv_soyad,
        'adv_mail': adv_mail,
        'adv_tel': adv_tel,
        'adv_adres': adv_adres,
        'adv_dogum_yeri': adv_dogum_yeri,
        'adv_nufus_il': adv_nufus_il,
        'adv_cinsiyet': adv_cinsiyet,
        'adv_ilce': adv_ilce,
        'sicil_min': sicil_min,
        'sicil_max': sicil_max,
        'order_by': order_by,
        'tag_filter': tag_filter,
        'stats': stats,
    })


_KARA_LISTE_SICIL = '_KARALIST_'
_BEYAZ_LISTE_SICIL = '_BEYAZLIST_'


def _sync_tag_to_lawyer_list(baro_lawyer, new_tag_type: str) -> int:
    """
    Etiket eklendiğinde/değiştiğinde/kaldırıldığında özel sistem avukat listelerini günceller.
    Kara Liste → "_KARALIST_" avukatı, cevap_status=olumsuz
    Beyaz Liste → "_BEYAZLIST_" avukatı, cevap_status=olumlu
    Returns: updated / added count
    """
    olumsuz, _ = StatusOption.objects.get_or_create(key='olumsuz', defaults={'label': 'Olumsuz'})
    olumlu, _ = StatusOption.objects.get_or_create(key='olumlu', defaults={'label': 'Olumlu'})

    # Person kaydını bul veya oluştur
    person = Person.objects.filter(kisi_sicilno=baro_lawyer.sicil_no).first()
    if not person:
        person = Person.objects.create(
            kisi_sicilno=baro_lawyer.sicil_no,
            ad=baro_lawyer.ad or baro_lawyer.sicil_no,
            soyad=baro_lawyer.soyad or '',
        )

    count = 0
    if new_tag_type == BaroLawyerTag.BLACKLIST:
        kara_lawyer, _ = Lawyer.objects.get_or_create(
            sicil_no=_KARA_LISTE_SICIL,
            defaults={'ad': 'Kara Liste', 'soyad': 'Sistemi'},
        )
        lp, created = LawyerPerson.objects.get_or_create(
            lawyer=kara_lawyer,
            kisi_sicilno=baro_lawyer.sicil_no,
            defaults={
                'person': person,
                'ad': baro_lawyer.ad or '',
                'soyad': baro_lawyer.soyad or '',
                'cevap_status': olumsuz,
                'active': True,
            },
        )
        if not created:
            lp.active = True
            lp.cevap_status = olumsuz
            lp.save(update_fields=['active', 'cevap_status', 'updated_at'])
        count = 1
        # Beyaz Liste'den çıkar
        beyaz = Lawyer.objects.filter(sicil_no=_BEYAZ_LISTE_SICIL).first()
        if beyaz:
            LawyerPerson.objects.filter(lawyer=beyaz, kisi_sicilno=baro_lawyer.sicil_no).update(active=False)

    elif new_tag_type == BaroLawyerTag.WHITELIST:
        beyaz_lawyer, _ = Lawyer.objects.get_or_create(
            sicil_no=_BEYAZ_LISTE_SICIL,
            defaults={'ad': 'Beyaz Liste', 'soyad': 'Sistemi'},
        )
        lp, created = LawyerPerson.objects.get_or_create(
            lawyer=beyaz_lawyer,
            kisi_sicilno=baro_lawyer.sicil_no,
            defaults={
                'person': person,
                'ad': baro_lawyer.ad or '',
                'soyad': baro_lawyer.soyad or '',
                'cevap_status': olumlu,
                'active': True,
            },
        )
        if not created:
            lp.active = True
            lp.cevap_status = olumlu
            lp.save(update_fields=['active', 'cevap_status', 'updated_at'])
        count = 1
        # Kara Liste'den çıkar
        kara = Lawyer.objects.filter(sicil_no=_KARA_LISTE_SICIL).first()
        if kara:
            LawyerPerson.objects.filter(lawyer=kara, kisi_sicilno=baro_lawyer.sicil_no).update(active=False)

    else:
        # Etiket kaldırıldı — her iki listeden de sil
        for sicil in [_KARA_LISTE_SICIL, _BEYAZ_LISTE_SICIL]:
            lawyer = Lawyer.objects.filter(sicil_no=sicil).first()
            if lawyer:
                c = LawyerPerson.objects.filter(
                    lawyer=lawyer, kisi_sicilno=baro_lawyer.sicil_no
                ).update(active=False)
                count += c

    return count


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_tagged_page(request, tag_type: str):
    """
    Kara Liste veya Beyaz Liste sayfası — etiketli tüm Baro kayıtları.
    tag_type: 'blacklist' | 'whitelist'
    """
    if tag_type not in ('blacklist', 'whitelist'):
        return redirect('ui_baro_lawyers')

    q = (request.GET.get('q') or '').strip()
    ilce = (request.GET.get('ilce') or '').strip()

    tags_qs = (
        BaroLawyerTag.objects
        .filter(tag_type=tag_type)
        .select_related('baro_lawyer')
        .order_by('baro_lawyer__sicil_no')
    )

    if q:
        tags_qs = tags_qs.filter(
            Q(baro_lawyer__ad__icontains=q) |
            Q(baro_lawyer__soyad__icontains=q) |
            Q(baro_lawyer__sicil_no__icontains=q)
        )
    if ilce:
        tags_qs = tags_qs.filter(baro_lawyer__ilce__iexact=ilce)

    tags = tags_qs

    # Her kişi için LawyerPerson durumunu bul (sistem avukatları hariç)
    sicil_list = [t.baro_lawyer.sicil_no for t in tags]
    from collections import defaultdict
    lp_qs = (
        LawyerPerson.objects
        .filter(kisi_sicilno__in=sicil_list, active=True)
        .exclude(lawyer__sicil_no__in=[_KARA_LISTE_SICIL, _BEYAZ_LISTE_SICIL])
        .select_related('cevap_status', 'lawyer')
        .values('kisi_sicilno', 'cevap_status__key', 'cevap_status__label', 'lawyer__ad', 'lawyer__soyad')
    )
    lp_map = defaultdict(list)
    for lp in lp_qs:
        lp_map[lp['kisi_sicilno']].append({
            'status_key': lp['cevap_status__key'] or '',
            'status_label': lp['cevap_status__label'] or '',
            'lawyer': f"{lp['lawyer__ad']} {lp['lawyer__soyad']}",
        })

    rows = []
    for tag in tags:
        rows.append({
            'tag': tag,
            'bl': tag.baro_lawyer,
            'lists': lp_map.get(tag.baro_lawyer.sicil_no, []),
        })

    label = 'Kara Liste' if tag_type == 'blacklist' else 'Beyaz Liste'
    return render(request, 'app/baro_tagged_list.html', {
        'rows': rows,
        'tag_type': tag_type,
        'label': label,
        'q': q,
        'ilce': ilce,
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_tagged_export(request, tag_type: str):
    """
    Kara Liste veya Beyaz Liste kayıtlarını CSV ya da Excel olarak indirir.
    GET params: format=csv|excel  q=arama  ilce=filtre
    """
    if tag_type not in ('blacklist', 'whitelist'):
        return redirect('ui_baro_lawyers')

    from datetime import datetime as _dt
    from io import BytesIO

    fmt = request.GET.get('format', 'csv')
    q = (request.GET.get('q') or '').strip()
    ilce_filter = (request.GET.get('ilce') or '').strip()

    qs = (
        BaroLawyerTag.objects
        .filter(tag_type=tag_type)
        .select_related('baro_lawyer')
        .order_by('baro_lawyer__sicil_no')
    )

    # İsim / sicil arama
    if q:
        qs = qs.filter(
            Q(baro_lawyer__ad__icontains=q) |
            Q(baro_lawyer__soyad__icontains=q) |
            Q(baro_lawyer__sicil_no__icontains=q)
        )

    # İlçe filtresi
    if ilce_filter:
        qs = qs.filter(baro_lawyer__ilce__iexact=ilce_filter)

    label = 'Kara Liste' if tag_type == 'blacklist' else 'Beyaz Liste'
    timestamp = _dt.now().strftime('%Y%m%d_%H%M%S')
    safe_label = 'kara_liste' if tag_type == 'blacklist' else 'beyaz_liste'

    COLUMNS = [
        ('sicil_no',     'Sicil No'),
        ('ad',           'Ad'),
        ('soyad',        'Soyad'),
        ('mail',         'E-posta'),
        ('tel',          'Telefon'),
        ('ilce',         'İlçe (Çalışma)'),
        ('dogum_tarihi', 'Doğum Tarihi'),
        ('uye',          'Üye Durumu'),
        ('tag_note',     'Etiket Notu'),
        ('created_by',   'Ekleyen'),
        ('created_at',   'Etiketlenme Tarihi'),
    ]

    def row_data(tag):
        bl = tag.baro_lawyer
        return [
            bl.sicil_no,
            bl.ad,
            bl.soyad,
            bl.mail or '',
            bl.tel or '',
            bl.ilce or '',
            bl.dogum_tarihi or '',
            bl.uye or '',
            tag.note or '',
            tag.created_by or '',
            tag.created_at.strftime('%d.%m.%Y') if tag.created_at else '',
        ]

    if fmt == 'excel':
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = label

        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        if tag_type == 'blacklist':
            header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        else:
            header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),  bottom=Side(style='thin', color='CCCCCC'),
        )
        col_widths = [12, 18, 18, 28, 16, 18, 14, 12, 35, 20, 16]

        # Bilgi satırı
        ws.merge_cells(f'A1:{get_column_letter(len(COLUMNS))}1')
        ws['A1'].value = f'{label} — Export: {_dt.now().strftime("%d.%m.%Y %H:%M")} — Toplam: {qs.count()} kayıt'
        ws['A1'].font = Font(name='Calibri', size=10, italic=True, color='666666')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Başlıklar
        for ci, (_, col_label) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=2, column=ci, value=col_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]

        # Veriler
        for ri, tag in enumerate(qs, 3):
            for ci, val in enumerate(row_data(tag), 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(name='Calibri', size=10)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top')

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:{get_column_letter(len(COLUMNS))}{qs.count() + 2}'

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        response = HttpResponse(
            bio.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{safe_label}_{timestamp}.xlsx"'
        return response

    else:  # csv
        import csv as _csv
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{safe_label}_{timestamp}.csv"'
        writer = _csv.writer(response)
        writer.writerow([col_label for _, col_label in COLUMNS])
        for tag in qs:
            writer.writerow(row_data(tag))
        return response


@login_required_custom
@require_http_methods(["POST"])
def ui_baro_tag_toggle(request, sicil_no: str):
    """
    Baro kaydına kara liste / 'biz' etiketi ekle veya kaldır.
    POST body: tag_type ('blacklist' | 'whitelist' | '' = kaldır), note (opsiyonel)
    """
    try:
        baro_lawyer = BaroLawyer.objects.get(sicil_no=sicil_no)
    except BaroLawyer.DoesNotExist:
        return JsonResponse({'error': 'Baro kaydı bulunamadı'}, status=404)

    tag_type = (request.POST.get('tag_type') or '').strip()
    note = (request.POST.get('note') or '').strip() or None
    actor = str(request.user) if request.user.is_authenticated else None

    if not tag_type:
        # Etiketi kaldır
        BaroLawyerTag.objects.filter(baro_lawyer=baro_lawyer).delete()
        _sync_tag_to_lawyer_list(baro_lawyer, '')
        return JsonResponse({'ok': True, 'action': 'removed', 'tag_type': None, 'status_updated': 0})

    if tag_type not in (BaroLawyerTag.BLACKLIST, BaroLawyerTag.WHITELIST):
        return JsonResponse({'error': 'Geçersiz etiket türü'}, status=400)

    tag, created = BaroLawyerTag.objects.update_or_create(
        baro_lawyer=baro_lawyer,
        defaults={'tag_type': tag_type, 'note': note, 'created_by': actor},
    )

    # Sistem avukat listesine ekle (kara/beyaz liste sanal avukatı)
    _sync_tag_to_lawyer_list(baro_lawyer, tag_type)

    # Diğer avukat listelerinde de cevap durumunu güncelle
    target_key = 'olumlu' if tag_type == BaroLawyerTag.WHITELIST else 'olumsuz'
    status_obj, _ = StatusOption.objects.get_or_create(
        key=target_key, defaults={'label': 'Olumlu' if target_key == 'olumlu' else 'Olumsuz'}
    )
    status_updated = LawyerPerson.objects.filter(
        kisi_sicilno=baro_lawyer.sicil_no, active=True
    ).update(cevap_status=status_obj)

    return JsonResponse({
        'ok': True,
        'action': 'created' if created else 'updated',
        'tag_type': tag_type,
        'note': tag.note or '',
        'status_updated': status_updated,
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_tag_stats(request):
    """Blacklist/whitelist istatistikleri — dashboard grafik için."""
    total = BaroLawyer.objects.count()
    blacklist_count = BaroLawyerTag.objects.filter(tag_type='blacklist').count()
    whitelist_count = BaroLawyerTag.objects.filter(tag_type='whitelist').count()
    return JsonResponse({
        'total': total,
        'blacklist': blacklist_count,
        'whitelist': whitelist_count,
        'untagged': total - blacklist_count - whitelist_count,
    })


def _tag_detail_response(tag_type: str):
    """Ortak yardımcı: belirli tag türündeki kişileri JSON olarak döndür."""
    tags = BaroLawyerTag.objects.filter(
        tag_type=tag_type
    ).select_related('baro_lawyer').order_by('baro_lawyer__sicil_no')
    data = [{
        'sicil_no': t.baro_lawyer.sicil_no,
        'ad': t.baro_lawyer.ad,
        'soyad': t.baro_lawyer.soyad,
        'mail': t.baro_lawyer.mail,
        'tel': t.baro_lawyer.tel,
        'note': t.note or '',
        'created_by': t.created_by or '',
        'created_at': t.created_at.strftime('%d.%m.%Y %H:%M'),
    } for t in tags]
    return JsonResponse({'ok': True, 'count': len(data), 'results': data})


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_blacklist_detail(request):
    """Kara listedeki kişilerin detay listesi (JSON)."""
    return _tag_detail_response('blacklist')


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_whitelist_detail(request):
    """Beyaz listedeki kişilerin detay listesi (JSON)."""
    return _tag_detail_response('whitelist')


@login_required_custom
@require_http_methods(["POST"])
def ui_baro_bulk_tag(request):
    """
    Toplu etiketleme — Birden fazla sicil no'yu aynı anda işaretle.
    POST body: tag_type ('blacklist' | 'whitelist'), sicil_nos (newline-separated), note (optional)
    """
    tag_type = (request.POST.get('tag_type') or '').strip()
    note = (request.POST.get('note') or '').strip() or None
    raw_sicils = request.POST.get('sicil_nos', '')
    actor = str(request.user) if request.user.is_authenticated else None

    if tag_type not in (BaroLawyerTag.BLACKLIST, BaroLawyerTag.WHITELIST):
        return JsonResponse({'error': 'Geçersiz etiket türü'}, status=400)

    sicil_list = [s.strip() for s in raw_sicils.replace(',', '\n').splitlines() if s.strip()]
    if not sicil_list:
        return JsonResponse({'error': 'Sicil no listesi boş'}, status=400)

    # Etiket türüne göre cevap durumu nesnesini hazırla
    target_key = 'olumlu' if tag_type == BaroLawyerTag.WHITELIST else 'olumsuz'
    target_label = 'Olumlu' if target_key == 'olumlu' else 'Olumsuz'
    status_obj, _ = StatusOption.objects.get_or_create(
        key=target_key, defaults={'label': target_label}
    )

    found, not_found, updated = [], [], 0
    total_status_updated = 0
    for sicil in sicil_list:
        try:
            baro_lawyer = BaroLawyer.objects.get(sicil_no=sicil)
            BaroLawyerTag.objects.update_or_create(
                baro_lawyer=baro_lawyer,
                defaults={'tag_type': tag_type, 'note': note, 'created_by': actor},
            )
            # Sistem avukat listesine ekle
            _sync_tag_to_lawyer_list(baro_lawyer, tag_type)
            # Diğer listelerdeki cevap durumunu da güncelle
            cnt = LawyerPerson.objects.filter(kisi_sicilno=sicil, active=True).update(cevap_status=status_obj)
            total_status_updated += cnt
            found.append(sicil)
            updated += 1
        except BaroLawyer.DoesNotExist:
            not_found.append(sicil)

    return JsonResponse({
        'ok': True,
        'updated': updated,
        'status_updated': total_status_updated,
        'not_found': not_found,
        'not_found_count': len(not_found),
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_lawyer_detail(request, sicil_no: str):
    """Baro avukatının tüm alanlarını JSON olarak döner (detay modal için)."""
    try:
        bl = BaroLawyer.objects.select_related('tag').get(sicil_no=sicil_no)
    except BaroLawyer.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'Bulunamadı'}, status=404)

    tag = getattr(bl, 'tag', None)
    memberships = list(bl.memberships.values_list('gorev', flat=True))

    photo_url = bl.photo_url or ''

    return JsonResponse({
        'ok': True,
        'sicil_no': bl.sicil_no,
        'ad': bl.ad,
        'soyad': bl.soyad,
        'tel': bl.tel,
        'mail': bl.mail,
        'adres': bl.adres,
        'uye': bl.uye,
        'cinsiyet': bl.cinsiyet,
        'mesleğe_baslama': getattr(bl, 'mesleğe_baslama', ''),
        'ilce': bl.ilce,
        'dogum_yeri': bl.dogum_yeri,
        'dogum_tarihi': bl.dogum_tarihi,
        'mahalle_koy': bl.mahalle_koy,
        'nufus_ilce': bl.nufus_ilce,
        'nufus_il': bl.nufus_il,
        'tag_type': tag.tag_type if tag else '',
        'tag_note': tag.note if tag else '',
        'memberships': memberships,
        'photo_url': photo_url,
        'has_photo': bool(photo_url),
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_baro_photo_proxy(request, sicil_no: str):
    """Baro fotoğrafını BARO_SESSION_COOKIE ile çekip akıtır (Railway filesystem yok)."""
    from django.conf import settings as _settings
    from django.http import StreamingHttpResponse

    try:
        bl = BaroLawyer.objects.only('photo_url').get(sicil_no=sicil_no)
    except BaroLawyer.DoesNotExist:
        return HttpResponse(status=404)

    photo_url = bl.photo_url or ''
    if not photo_url:
        return HttpResponse(status=404)

    session_cookie = getattr(_settings, 'BARO_SESSION_COOKIE', '')
    if not session_cookie:
        return HttpResponse(status=404)

    try:
        import requests as req
    except ImportError:
        return HttpResponse(status=503)

    try:
        resp = req.get(
            photo_url,
            headers={
                'cookie': session_cookie,
                'referer': 'https://www.ankarabarosu.org.tr/avukatlar/',
                'user-agent': (
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                    'AppleWebKit/537.36 (KHTML, like Gecko) '
                    'Chrome/146.0.0.0 Safari/537.36'
                ),
                'accept': 'image/avif,image/webp,image/apng,image/*,*/*;q=0.8',
                'sec-fetch-dest': 'image',
                'sec-fetch-mode': 'no-cors',
                'sec-fetch-site': 'same-origin',
            },
            stream=True,
            timeout=15,
        )
    except Exception:
        return HttpResponse(status=502)

    if resp.status_code != 200:
        return HttpResponse(status=resp.status_code)

    content_type = resp.headers.get('content-type', 'image/jpeg')
    streaming = StreamingHttpResponse(
        resp.iter_content(chunk_size=8192),
        content_type=content_type,
    )
    streaming['Cache-Control'] = 'private, max-age=86400'
    return streaming


@login_required_custom
@require_http_methods(["GET"])
def ui_birthdays_today(request):
    """Bugün doğum günü olan baro avukatlarını döner (liste ataması bilgisiyle)"""
    from django.utils import timezone as _tz
    today = _tz.localdate()  # Django TIME_ZONE ayarını kullan (Europe/Istanbul)
    suffix = f"-{today.month:02d}-{today.day:02d}"
    lawyers_qs = BaroLawyer.objects.filter(
        dogum_tarihi__endswith=suffix
    ).exclude(dogum_tarihi='').values(
        'sicil_no', 'ad', 'soyad', 'dogum_tarihi', 'tel', 'mail'
    ).order_by('ad', 'soyad')

    lawyers_list = list(lawyers_qs)
    birthday_sicils = [l['sicil_no'] for l in lawyers_list]

    # Her kişi için en güncel LawyerPerson bilgisini al
    lp_map = {}
    for lp in LawyerPerson.objects.filter(
        kisi_sicilno__in=birthday_sicils, active=True
    ).select_related('cevap_status', 'lawyer').order_by('kisi_sicilno', '-id'):
        if lp.kisi_sicilno not in lp_map:
            lp_map[lp.kisi_sicilno] = {
                'in_list': True,
                'status': lp.cevap_status.key if lp.cevap_status else None,
                'status_label': lp.cevap_status.label if lp.cevap_status else None,
                'lawyer_name': f"{lp.lawyer.ad} {lp.lawyer.soyad}",
            }

    # Kara/beyaz liste bilgisi
    tag_map = {}
    for tag in BaroLawyerTag.objects.filter(
        baro_lawyer__sicil_no__in=birthday_sicils
    ).select_related('baro_lawyer'):
        tag_map[tag.baro_lawyer.sicil_no] = tag.tag_type

    for l in lawyers_list:
        info = lp_map.get(l['sicil_no'], {
            'in_list': False, 'status': None, 'status_label': None, 'lawyer_name': None,
        })
        l.update(info)
        l['tag_type'] = tag_map.get(l['sicil_no'])

    return JsonResponse({
        'ok': True,
        'date': today.isoformat(),
        'count': len(lawyers_list),
        'lawyers': lawyers_list,
    })


@login_required_custom
@require_http_methods(["GET"])
def ui_committee_memberships(request):
    """
    Kurul üyeliklerini listeler - Baro_Merkezler_vs.xlsx verisinden
    """
    q = (request.GET.get('q') or '').strip()
    adv_sicil = (request.GET.get('sicil') or '').strip()
    adv_ad_soyad = (request.GET.get('ad_soyad') or '').strip()
    adv_gorev = (request.GET.get('gorev') or '').strip()
    order_by = request.GET.get('order_by', 'ad_soyad').strip()

    qs = CommitteeMembership.objects.select_related('baro_lawyer').all()

    if q:
        qs = qs.filter(
            Q(ad_soyad__icontains=q) |
            Q(sicil_no__icontains=q) |
            Q(gorev__icontains=q)
        )
    if adv_sicil:
        qs = qs.filter(sicil_no__icontains=adv_sicil)
    if adv_ad_soyad:
        qs = qs.filter(ad_soyad__icontains=adv_ad_soyad)
    if adv_gorev:
        qs = qs.filter(gorev__icontains=adv_gorev)

    allowed_orders = ['ad_soyad', '-ad_soyad', 'sicil_no', '-sicil_no', 'gorev', '-gorev']
    if order_by not in allowed_orders:
        order_by = 'ad_soyad'
    qs = qs.order_by(order_by)

    # İstatistikler
    from django.db.models import Count
    total = CommitteeMembership.objects.count()
    gorev_list = CommitteeMembership.objects.values('gorev').annotate(count=Count('id')).order_by('gorev')

    page = Paginator(qs, 25).get_page(request.GET.get('page'))

    return render(request, 'app/committee_memberships.html', {
        'page': page,
        'q': q,
        'adv_sicil': adv_sicil,
        'adv_ad_soyad': adv_ad_soyad,
        'adv_gorev': adv_gorev,
        'order_by': order_by,
        'total': total,
        'gorev_list': gorev_list,
    })
